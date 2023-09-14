# chao
# 时间：2023/9/13 4:37 下午

#import openpyxl
from openpyxl import load_workbook

workbook = load_workbook('data.xlsx')  #读取excel表格
worksheet = workbook.worksheets[0]  #读取sheet页

data_list =  []  #创建一个空的列表

# 循环遍历列数据，存储到列表
for row in worksheet.iter_rows(values_only=True):
    data_list.append(row[0])
print(data_list)